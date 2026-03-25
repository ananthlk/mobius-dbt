#!/usr/bin/env python3
"""
Generate Provider Roster / Credentialing report for an organization.
Uses the provider-roster-credentialing skill core (same logic as API).

Usage:
  uv run python scripts/generate_provider_roster_credentialing_report.py --org-name "David Lawrence"
  uv run python scripts/generate_provider_roster_credentialing_report.py --org-name "David Lawrence" --output-dir reports/
  uv run python scripts/generate_provider_roster_credentialing_report.py --org-name "David Lawrence" --enhance
  uv run python scripts/generate_provider_roster_credentialing_report.py --org-name "Aspire" --locations locations.json --npi-overrides npi_overrides.json --output-dir reports/

Output:
  Provider Roster / Credentialing report: executive summary (MD), locations.csv, npis_per_location.csv,
  per_npi_validation.csv, combos.csv, invalid_combos.csv, ghost_billing.csv; optional XLSX with sheets.
  With --enhance: main MD is LLM-generated white-paper (overview, methodology, findings, insights, sources);
  raw metrics are written to prefix_raw_summary.md.

Env:
  BQ_PROJECT, BQ_MARTS_MEDICAID_DATASET, BQ_LANDING_MEDICAID_DATASET
  For --enhance: OPENAI_API_KEY or GEMINI_API_KEY / VERTEX_PROJECT_ID
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from datetime import datetime
from pathlib import Path

# Add skill package to path so we can import app.core
_repo_root = Path(__file__).resolve().parents[2]  # mobius-dbt/scripts -> Mobius repo root
_skill_path = _repo_root / "mobius-skills" / "provider-roster-credentialing"
if _skill_path.is_dir():
    sys.path.insert(0, str(_skill_path))
else:
    sys.path.insert(0, str(_repo_root))


def _fixture_report(org_name: str) -> dict:
    """Fallback report when BigQuery access fails; used for white-paper preview."""
    return {
        "executive_summary": {
            "org_name": org_name,
            "location_count": 4,
            "total_npis": 188,
            "npis_all_checks_pass": 124,
            "npis_at_least_one_fail": 64,
            "invalid_combo_count": 247,
            "ghost_billing_npi_count": 0,
            "ghost_billing_claim_count": 0,
            "ghost_billing_total_paid": 0,
            "readiness_status_breakdown": {
                "Ready": 612,
                "Not enrolled": 117,
                "Combo mismatch": 69,
                "Taxonomy not permitted": 8,
                "Invalid address": 53,
            },
            "next_steps": "247 invalid combo(s) need resolution.",
            "revenue_at_risk_2024": 0,
            "billing_impact_note": None,
            "revenue_at_risk_2024_by_status": {},
            "confidence_breakdown": {"high": 0, "medium": 0, "low": 0},
            "readiness_score": 71,
        },
        "locations": [
            {"org_name": org_name, "site_city": "Naples", "site_state": "FL", "site_zip": "34102"},
            {"org_name": org_name, "site_city": "Fort Myers", "site_state": "FL", "site_zip": "33901"},
            {"org_name": org_name, "site_city": "Bonita Springs", "site_state": "FL", "site_zip": "34134"},
            {"org_name": org_name, "site_city": "Fort Myers", "site_state": "FL", "site_zip": "33916"},
        ],
        "npis_per_location": {},
        "per_npi_validation": [],
        "combos": [],
        "invalid_combos": [
            {"servicing_npi": "1234567890", "servicing_provider_name": "Sample", "readiness_status": "Not enrolled", "readiness_summary": "Fixture data."},
        ],
        "missed_opportunities": [],
        "ghost_billing": [],
    }


def _org_folder_name(org_name: str) -> str:
    """Sanitize org name for use as folder name (e.g. 'Aspire Health' -> 'Aspire_Health')."""
    import re
    s = org_name.strip()
    s = re.sub(r'[^\w\s-]', '', s)
    s = re.sub(r'[\s-]+', '_', s)
    return s or "org"


def _load_json_path(path: str | None) -> dict | list | None:
    if not path or not path.strip():
        return None
    p = Path(path.strip())
    if not p.is_file():
        return None
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def _write_csv(filepath: Path, rows: list[dict], fieldnames: list[str] | None = None) -> None:
    if not rows:
        return
    keys = fieldnames or list(rows[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if v is None else str(v)) for k, v in r.items()})


def _write_md(filepath: Path, report: dict) -> None:
    ex = report.get("executive_summary") or {}
    if ex.get("error"):
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("# Provider Roster / Credentialing Report\n\n")
            f.write(f"**Error:** {ex['error']}\n")
        return
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("# Provider Roster / Credentialing Report\n\n")
        f.write(f"**Organization:** {ex.get('org_name', '')}  \n")
        f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n---\n\n")
        f.write("## Executive Summary\n\n")
        f.write("| Metric | Value |\n|--------|------|\n")
        f.write(f"| Locations | {ex.get('location_count', 0)} |\n")
        f.write(f"| Total NPIs | {ex.get('total_npis', 0)} |\n")
        f.write(f"| NPIs (all checks pass) | {ex.get('npis_all_checks_pass', 0)} |\n")
        f.write(f"| NPIs (at least one fail) | {ex.get('npis_at_least_one_fail', 0)} |\n")
        f.write(f"| Invalid combos | {ex.get('invalid_combo_count', 0)} |\n")
        f.write(f"| Ghost billing (NPIs) | {ex.get('ghost_billing_npi_count', 0)} |\n")
        f.write(f"| Ghost billing (claims) | {ex.get('ghost_billing_claim_count', 0)} |\n")
        f.write(f"| Ghost billing ($) | ${ex.get('ghost_billing_total_paid', 0):,.0f} |\n")
        f.write("\n**Readiness status breakdown:**\n\n")
        for status, count in (ex.get("readiness_status_breakdown") or {}).items():
            f.write(f"- {status}: {count}\n")
        f.write(f"\n**Next steps:** {ex.get('next_steps', '')}\n\n---\n\n")
        f.write("## Outputs\n\n")
        f.write("See CSV files in the same directory: locations, npis_per_location, per_npi_validation, combos, invalid_combos, ghost_billing.\n")


def main() -> int:
    try:
        _r, _dbt = Path(__file__).resolve().parents[2], Path(__file__).resolve().parents[1]
        if (_r / "mobius-config").exists():
            sys.path.insert(0, str(_r / "mobius-config"))
            from env_helper import load_env
            load_env(_dbt)
    except Exception:
        pass
    parser = argparse.ArgumentParser(
        description="Generate Provider Roster / Credentialing report for an organization."
    )
    parser.add_argument("--org-name", required=True, help='Organization name (e.g. "David Lawrence")')
    parser.add_argument("--locations", default=None, help="Optional JSON file: list of location_id to include, or dict with location_ids key")
    parser.add_argument("--locations-override", default=None, help="Optional JSON file: L2 user-validated locations. List of {site_address_line_1, site_city, site_state, site_zip} or dict with locations_override key. Replaces system-imputed locations (universal truth).")
    parser.add_argument("--npi-overrides", default=None, help="Optional JSON file: dict location_id -> { add: [npi,...], remove: [npi,...] }")
    parser.add_argument("--output-dir", default=None, help="Output directory (default: mobius-dbt/reports)")
    parser.add_argument("--state", default="FL", help="Filter locations to this state (default: FL). Avoids name collision (e.g. Henderson FL vs Henderson NV).")
    parser.add_argument("--enhance", action="store_true", help="Use LLM to generate white-paper-style report. Uses VERTEX_PROJECT_ID or CHAT_VERTEX_PROJECT_ID (same as other modules) or OPENAI_API_KEY / GEMINI_API_KEY.")
    parser.add_argument("--no-pipeline", action="store_true", help="With --enhance: skip Validator/Critic/Composer pipeline (single Drafter only). Default: use full pipeline.")
    parser.add_argument("--no-pdf", action="store_true", help="With --enhance: skip PDF generation (PDF is default for shareable reports).")
    parser.add_argument("--llm-provider", choices=("openai", "gemini"), default=None, help="LLM provider when using --enhance.")
    parser.add_argument("--llm-model", default=None, help="Model when using --enhance (e.g. gpt-4o, gemini-1.5-pro).")
    args = parser.parse_args()

    try:
        from google.cloud import bigquery
    except ImportError:
        print("Install: pip install google-cloud-bigquery", file=sys.stderr)
        return 1

    try:
        from app.core import build_full_report
    except ImportError:
        print("Cannot import app.core. Ensure mobius-skills/provider-roster-credentialing exists.", file=sys.stderr)
        return 1

    project = os.environ.get("BQ_PROJECT", "mobius-os-dev")
    marts_dataset = os.environ.get("BQ_MARTS_MEDICAID_DATASET", "mobius_medicaid_npi_dev")
    landing_dataset = os.environ.get("BQ_LANDING_MEDICAID_DATASET", "landing_medicaid_npi_dev")
    base_reports = Path(args.output_dir) if args.output_dir else Path(__file__).resolve().parents[1] / "reports"
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    org_folder = _org_folder_name(args.org_name.strip())
    output_dir = base_reports / org_folder / ts
    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"Output: {output_dir}")

    location_ids = None
    loc_json = _load_json_path(args.locations)
    if loc_json is not None:
        if isinstance(loc_json, list):
            location_ids = [str(x) for x in loc_json]
        elif isinstance(loc_json, dict) and "location_ids" in loc_json:
            location_ids = [str(x) for x in loc_json["location_ids"]]

    locations_override = None
    loc_override_json = _load_json_path(getattr(args, "locations_override", None))
    if loc_override_json is not None:
        if isinstance(loc_override_json, list):
            locations_override = loc_override_json
        elif isinstance(loc_override_json, dict) and "locations_override" in loc_override_json:
            locations_override = loc_override_json["locations_override"]
        if locations_override:
            print(f"  Using {len(locations_override)} user-validated locations (L2 override)")

    npi_overrides = None
    npi_json = _load_json_path(args.npi_overrides)
    if isinstance(npi_json, dict):
        npi_overrides = npi_json

    print(f"Building Provider Roster / Credentialing report for org_name={args.org_name!r}...")
    report = None
    try:
        client = bigquery.Client(project=project)
        report = build_full_report(
            client,
            org_name=args.org_name.strip(),
            project=project,
            marts_dataset=marts_dataset,
            landing_dataset=landing_dataset,
            location_ids=location_ids,
            npi_overrides=npi_overrides,
            locations_override=locations_override,
            state_filter=getattr(args, "state", "FL") or "FL",
        )
    except Exception as e:
        err_str = str(e).lower()
        if "403" in err_str or "forbidden" in err_str or "access denied" in err_str:
            if getattr(args, "enhance", False):
                print("BigQuery access denied (403). Using fixture data to generate white-paper preview.", file=sys.stderr)
                print("To fix BigQuery: 1) Enable BigQuery API for the project. 2) Grant your account "
                      "'BigQuery Job User' and 'BigQuery Data Viewer' (or 'BigQuery User') on the project. "
                      "3) If using gcloud: gcloud auth application-default login", file=sys.stderr)
                report = _fixture_report(args.org_name.strip())
            else:
                raise
        else:
            raise

    ex = report.get("executive_summary") or {}
    if ex.get("error"):
        print(f"Error: {ex['error']}", file=sys.stderr)
        _write_md(output_dir / "provider_roster_credentialing_report.md", report)
        return 1

    prefix = f"provider_roster_credentialing_{ts}"

    # npis_per_location: flatten for CSV
    npi_rows = []
    for loc_id, nlist in (report.get("npis_per_location") or {}).items():
        for n in nlist:
            npi_rows.append({"location_id": loc_id, **n})

    # Primary report first (simple sheet: locations, distinct NPIs + why, taxonomies, billing)
    primary = report.get("primary_report") or {}
    _write_csv(output_dir / f"{prefix}_primary_locations.csv", primary.get("locations") or [])
    _write_csv(output_dir / f"{prefix}_primary_distinct_npis.csv", primary.get("distinct_npis") or [])
    _write_csv(output_dir / f"{prefix}_primary_taxonomies_covered.csv", primary.get("taxonomies_covered") or [])
    _write_csv(output_dir / f"{prefix}_primary_billing_activity.csv", primary.get("billing_activity") or [])
    print(f"  Wrote primary report (locations, distinct NPIs, taxonomies, billing)")

    # Rest of report
    _write_csv(output_dir / f"{prefix}_locations.csv", report.get("locations") or [])
    print(f"  Wrote {prefix}_locations.csv")
    _write_csv(output_dir / f"{prefix}_npis_per_location.csv", npi_rows)
    print(f"  Wrote {prefix}_npis_per_location.csv")

    _write_csv(output_dir / f"{prefix}_per_npi_validation.csv", report.get("per_npi_validation") or [])
    print(f"  Wrote {prefix}_per_npi_validation.csv")

    _write_csv(output_dir / f"{prefix}_combos.csv", report.get("combos") or [])
    print(f"  Wrote {prefix}_combos.csv")

    _write_csv(output_dir / f"{prefix}_confidence_report.csv", report.get("confidence_report") or [])
    print(f"  Wrote {prefix}_confidence_report.csv")

    _write_csv(output_dir / f"{prefix}_invalid_combos.csv", report.get("invalid_combos") or [])
    print(f"  Wrote {prefix}_invalid_combos.csv")

    _write_csv(output_dir / f"{prefix}_ghost_billing.csv", report.get("ghost_billing") or [])
    print(f"  Wrote {prefix}_ghost_billing.csv")

    _write_csv(output_dir / f"{prefix}_missed_opportunities.csv", report.get("missed_opportunities") or [])
    print(f"  Wrote {prefix}_missed_opportunities.csv")

    _write_csv(output_dir / f"{prefix}_locations_match_report.csv", report.get("locations_match_report") or [])
    print(f"  Wrote {prefix}_locations_match_report.csv")

    _write_csv(output_dir / f"{prefix}_l1_not_in_l2.csv", report.get("l1_not_in_l2") or [])
    print(f"  Wrote {prefix}_l1_not_in_l2.csv")

    # metrics.json for Data Validator (pipeline canonical metrics)
    ex = report.get("executive_summary") or {}
    status_breakdown = ex.get("readiness_status_breakdown") or {}
    ready_count = status_breakdown.get("Ready") or 0
    invalid_count = ex.get("invalid_combo_count") or 0
    total_combo_count = ready_count + invalid_count
    metrics = {
        "org_name": ex.get("org_name"),
        "location_count": ex.get("location_count", 0),
        "total_npis": ex.get("total_npis", 0),
        "npis_with_readiness": ex.get("npis_with_readiness", 0),
        "npis_no_readiness": ex.get("npis_no_readiness", 0),
        "npis_org_misaligned": ex.get("npis_org_misaligned", 0),
        "npis_all_checks_pass": ex.get("npis_all_checks_pass", 0),
        "npis_at_least_one_fail": ex.get("npis_at_least_one_fail", 0),
        "invalid_combo_count": invalid_count,
        "ready_combo_count": ready_count,
        "total_combo_count": total_combo_count,
        "readiness_status_breakdown": status_breakdown,
        "revenue_at_risk_2024": ex.get("revenue_at_risk_2024"),
        "revenue_at_risk_2024_low": ex.get("revenue_at_risk_2024_low"),
        "revenue_at_risk_2024_high": ex.get("revenue_at_risk_2024_high"),
        "deprecated_taxonomy_revenue": ex.get("deprecated_taxonomy_revenue"),
        "revenue_by_location": ex.get("revenue_by_location") or [],
        "revenue_by_taxonomy": ex.get("revenue_by_taxonomy") or [],
        "revenue_assumptions": ex.get("revenue_assumptions") or {},
        "top_recommendations_by_code": ex.get("top_recommendations_by_code") or [],
        "recommendations_by_problem_type": ex.get("recommendations_by_problem_type") or [],
        "opportunity_confidence_matrix": ex.get("opportunity_confidence_matrix") or [],
        "confidence_definitions": ex.get("confidence_definitions") or {},
        "revenue_at_risk_2024_by_status": ex.get("revenue_at_risk_2024_by_status") or {},
        "revenue_at_risk_2024_by_confidence": ex.get("revenue_at_risk_2024_by_confidence") or {},
        "confidence_breakdown": ex.get("confidence_breakdown") or {"high": 0, "medium": 0, "low": 0},
        "readiness_score": ex.get("readiness_score"),
        "ghost_billing_npi_count": ex.get("ghost_billing_npi_count", 0),
        "ghost_billing_claim_count": ex.get("ghost_billing_claim_count", 0),
        "ghost_billing_total_paid": ex.get("ghost_billing_total_paid", 0),
        "missed_opportunities_count": len(report.get("missed_opportunities") or []),
    }
    with open(output_dir / f"{prefix}_metrics.json", "w", encoding="utf-8") as f:
        json.dump(metrics, f, indent=2)
    print(f"  Wrote {prefix}_metrics.json")

    # Generate chart images (when --enhance)
    generated_charts = []
    if getattr(args, "enhance", False):
        try:
            from app.report_visuals import get_chart_spec_from_llm, generate_charts
            provider = getattr(args, "llm_provider", None) or ("openai" if os.getenv("OPENAI_API_KEY", "").strip() else None) or ("gemini" if (os.getenv("GEMINI_API_KEY", "").strip() or os.getenv("BQ_PROJECT", "").strip() or os.getenv("VERTEX_PROJECT_ID", "").strip() or os.getenv("CHAT_VERTEX_PROJECT_ID", "").strip()) else None)
            chart_ids = None
            if provider:
                chart_ids = get_chart_spec_from_llm(report, metrics, provider=provider, model=getattr(args, "llm_model", None))
            generated_charts = generate_charts(report, metrics, output_dir, prefix, chart_ids=chart_ids or None)
            if generated_charts:
                report["_generated_charts"] = generated_charts
                for c in generated_charts:
                    print(f"  Wrote {c['filename']}")
        except ImportError as e:
            print(f"  Chart generation skipped: {e}", file=sys.stderr)
        except Exception as e:
            print(f"  Chart generation failed: {e}", file=sys.stderr)

    # LLM enhance: pipeline (Drafter+Validator+Critic+Composer) or single Drafter
    if getattr(args, "enhance", False):
        use_pipeline = not getattr(args, "no_pipeline", False)
        try:
            from app.report_writer import generate_white_paper_report
            if use_pipeline:
                from app.report_pipeline import generate_with_pipeline
        except ImportError as e:
            print(f"Cannot import report modules: {e}", file=sys.stderr)
            return 1
        provider = getattr(args, "llm_provider", None) or ("openai" if os.getenv("OPENAI_API_KEY", "").strip() else None) or ("gemini" if (os.getenv("GEMINI_API_KEY", "").strip() or os.getenv("BQ_PROJECT", "").strip() or os.getenv("VERTEX_PROJECT_ID", "").strip() or os.getenv("CHAT_VERTEX_PROJECT_ID", "").strip()) else None)
        if not provider:
            print("For --enhance set OPENAI_API_KEY or GEMINI_API_KEY / VERTEX_PROJECT_ID, or pass --llm-provider openai|gemini.", file=sys.stderr)
            return 1
        if use_pipeline:
            print("Generating white-paper via pipeline (Drafter → Validator + Critic → Composer)...")
            csv_names = [
                "locations.csv", "npis_per_location.csv", "per_npi_validation.csv",
                "combos.csv", "confidence_report.csv", "invalid_combos.csv",
                "ghost_billing.csv", "missed_opportunities.csv",
            ]
            csv_contents = {}
            _max_csv_chars = 80000
            for name in csv_names:
                p = output_dir / f"{prefix}_{name}"
                if p.exists():
                    raw = p.read_text(encoding="utf-8")
                    if len(raw) > _max_csv_chars:
                        csv_contents[name] = raw[:_max_csv_chars] + f"\n\n[TRUNCATED - total {len(raw)} chars; use metrics.json for counts]"
                    else:
                        csv_contents[name] = raw
                else:
                    csv_contents[name] = ""
            try:
                final_md, validation_report, critique_report, draft_md = generate_with_pipeline(
                    report, csv_contents, metrics,
                    provider=provider, model=getattr(args, "llm_model", None),
                )
            except Exception as e:
                print(f"Pipeline failed: {e}", file=sys.stderr)
                return 1
            with open(output_dir / f"{prefix}_draft.md", "w", encoding="utf-8") as f:
                f.write(draft_md)
            print(f"  Wrote {prefix}_draft.md (Drafter)")
            with open(output_dir / f"{prefix}.md", "w", encoding="utf-8") as f:
                f.write(final_md)
            print(f"  Wrote {prefix}.md (white-paper)")
            if not getattr(args, "no_pdf", False):
                try:
                    from app.report_pdf import markdown_to_pdf
                    pdf_path = output_dir / f"{prefix}.pdf"
                    if markdown_to_pdf(output_dir / f"{prefix}.md", pdf_path):
                        print(f"  Wrote {prefix}.pdf")
                    else:
                        print("  PDF generation skipped (see stderr)", file=sys.stderr)
                except ImportError as e:
                    print(f"  PDF skipped: {e}", file=sys.stderr)
            with open(output_dir / f"{prefix}_validation_report.md", "w", encoding="utf-8") as f:
                f.write(validation_report)
            print(f"  Wrote {prefix}_validation_report.md")
            with open(output_dir / f"{prefix}_critique_report.md", "w", encoding="utf-8") as f:
                f.write(critique_report)
            print(f"  Wrote {prefix}_critique_report.md")
        else:
            print("Generating white-paper-style report via LLM (Drafter only)...")
            try:
                enhanced_md = generate_white_paper_report(report, provider=provider, model=getattr(args, "llm_model", None))
            except Exception as e:
                print(f"LLM report generation failed: {e}", file=sys.stderr)
                return 1
            with open(output_dir / f"{prefix}.md", "w", encoding="utf-8") as f:
                f.write(enhanced_md)
            print(f"  Wrote {prefix}.md (white-paper)")
            if not getattr(args, "no_pdf", False):
                try:
                    from app.report_pdf import markdown_to_pdf
                    pdf_path = output_dir / f"{prefix}.pdf"
                    if markdown_to_pdf(output_dir / f"{prefix}.md", pdf_path):
                        print(f"  Wrote {prefix}.pdf")
                    else:
                        print("  PDF generation skipped (see stderr)", file=sys.stderr)
                except ImportError as e:
                    print(f"  PDF skipped: {e}", file=sys.stderr)
        _write_md(output_dir / f"{prefix}_raw_summary.md", report)
        print(f"  Wrote {prefix}_raw_summary.md (raw metrics)")
    else:
        _write_md(output_dir / f"{prefix}.md", report)
        print(f"  Wrote {prefix}.md")

    # Optional XLSX
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_name, rows in [
            ("Locations", report.get("locations") or []),
            ("Locations_match_report", report.get("locations_match_report") or []),
            ("L1_not_in_L2", report.get("l1_not_in_l2") or []),
            ("NPIs_per_location", npi_rows),
            ("Per_NPI_validation", report.get("per_npi_validation") or []),
            ("Combos", report.get("combos") or []),
            ("Invalid_combos", report.get("invalid_combos") or []),
            ("Ghost_billing", report.get("ghost_billing") or []),
            ("Missed_opportunities", report.get("missed_opportunities") or []),
        ]:
            if not rows:
                continue
            ws = wb.create_sheet(title=sheet_name[:31])
            keys = list(rows[0].keys())
            for c, k in enumerate(keys, 1):
                ws.cell(row=1, column=c, value=k)
            for r_idx, row in enumerate(rows, 2):
                for c_idx, k in enumerate(keys, 1):
                    v = row.get(k)
                    ws.cell(row=r_idx, column=c_idx, value="" if v is None else str(v))
        wb.save(output_dir / f"{prefix}.xlsx")
        print(f"  Wrote {prefix}.xlsx")
    except ImportError:
        pass  # openpyxl optional

    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
